from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

HEADER_KEYWORDS = {
    "序号",
    "元件名称",
    "值",
    "元件描述",
    "元件封装",
    "每套数量",
    "型号",
    "器件型号",
    "物料型号",
    "厂家型号",
    "商品编号",
    "商城编号",
    "立创编号",
    "mpn",
    "part number",
    "comment",
    "description",
    "designator",
    "footprint",
    "quantity",
}


@dataclass(slots=True)
class FilePreview:
    path: Path
    headers: list[str]
    rows: list[list[Any]]
    sheet_names: list[str]
    active_sheet: str | None = None
    header_row: int = 1


@dataclass(slots=True)
class SearchRow:
    row_number: int
    keyword: str


def read_preview(
    file_path: str | Path,
    max_rows: int = 20,
    sheet_name: str | None = None,
) -> FilePreview:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return _read_xlsx_preview(path, max_rows, sheet_name)
    if suffix == ".csv":
        return _read_csv_preview(path, max_rows)

    raise ValueError(f"Unsupported file format: {suffix}")


def read_preview_and_search_rows(
    file_path: str | Path,
    search_column: str,
    max_rows: int = 20,
    sheet_name: str | None = None,
) -> tuple[FilePreview, list[SearchRow]]:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return _read_xlsx_preview_and_search_rows(path, search_column, max_rows, sheet_name)

    preview = read_preview(path, max_rows, sheet_name)
    column_index = resolve_column_index(preview.headers, search_column)
    return preview, read_search_rows(path, column_index, sheet_name, preview.header_row)


def detect_header_row(
    file_path: str | Path,
    sheet_name: str | None = None,
    scan_rows: int = 30,
) -> int:
    path = Path(file_path)
    if path.name.startswith("~$"):
        raise ValueError("This looks like an Excel lock file; choose the original BOM file")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
            rows = worksheet.iter_rows(max_row=scan_rows, values_only=True)
            return _score_header_rows(rows)
        finally:
            workbook.close()
    if suffix == ".csv":
        dataframe = _read_csv(path, header=None, nrows=scan_rows)
        return _score_header_rows(dataframe.itertuples(index=False, name=None))

    raise ValueError(f"Unsupported file format: {suffix}")


def resolve_column_index(headers: list[str], selector: str) -> int:
    normalized_selector = selector.strip()
    if not normalized_selector:
        raise ValueError("Search column is required")

    if normalized_selector.isdigit():
        index = int(normalized_selector) - 1
        if 0 <= index < len(headers):
            return index
        raise ValueError(f"Column number out of range: {selector}")

    lowered_headers = [header.lower() for header in headers]
    lowered_selector = normalized_selector.lower()
    if lowered_selector in lowered_headers:
        return lowered_headers.index(lowered_selector)

    raise ValueError(f"Column not found: {selector}")


def read_search_rows(
    file_path: str | Path,
    column_index: int,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> list[SearchRow]:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return _read_xlsx_search_rows(path, column_index, sheet_name, header_row)
    if suffix == ".csv":
        return _read_csv_search_rows(path, column_index, header_row)

    raise ValueError(f"Unsupported file format: {suffix}")


def _read_xlsx_preview(path: Path, max_rows: int, sheet_name: str | None) -> FilePreview:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        header_row = detect_header_row(path, sheet_name)
        rows = list(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row + max_rows,
                values_only=True,
            )
        )
        headers = [
            _format_header(value, index) for index, value in enumerate(rows[0] if rows else [])
        ]
        preview_rows = [list(row) for row in rows[1:]]
        return FilePreview(
            path,
            headers,
            preview_rows,
            workbook.sheetnames,
            worksheet.title,
            header_row,
        )
    finally:
        workbook.close()


def _read_xlsx_preview_and_search_rows(
    path: Path,
    search_column: str,
    max_rows: int,
    sheet_name: str | None,
) -> tuple[FilePreview, list[SearchRow]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        header_row = _score_header_rows(worksheet.iter_rows(max_row=30, values_only=True))
        rows = list(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row + max_rows,
                values_only=True,
            )
        )
        headers = [
            _format_header(value, index) for index, value in enumerate(rows[0] if rows else [])
        ]
        preview = FilePreview(
            path,
            headers,
            [list(row) for row in rows[1:]],
            workbook.sheetnames,
            worksheet.title,
            header_row,
        )
        column_index = resolve_column_index(headers, search_column)
        search_rows: list[SearchRow] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if column_index >= len(row):
                continue
            value = row[column_index]
            keyword = "" if value is None else str(value).strip()
            if keyword:
                search_rows.append(SearchRow(row_number=row_number, keyword=keyword))
        return preview, search_rows
    finally:
        workbook.close()


def _read_csv_preview(path: Path, max_rows: int) -> FilePreview:
    header_row = detect_header_row(path)
    dataframe = _read_csv(path, nrows=max_rows, header=header_row - 1)
    headers = [_format_header(column, index) for index, column in enumerate(dataframe.columns)]
    return FilePreview(path, headers, dataframe.values.tolist(), [], None, header_row)


def _read_xlsx_search_rows(
    path: Path,
    column_index: int,
    sheet_name: str | None,
    header_row: int | None,
) -> list[SearchRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        resolved_header_row = header_row or detect_header_row(path, sheet_name)
        search_rows: list[SearchRow] = []

        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_number <= resolved_header_row or column_index >= len(row):
                continue
            value = row[column_index]
            keyword = "" if value is None else str(value).strip()
            if keyword:
                search_rows.append(SearchRow(row_number=row_number, keyword=keyword))

        return search_rows
    finally:
        workbook.close()


def _read_csv_search_rows(
    path: Path,
    column_index: int,
    header_row: int | None,
) -> list[SearchRow]:
    resolved_header_row = header_row or detect_header_row(path)
    dataframe = _read_csv(path, header=resolved_header_row - 1)
    if column_index >= len(dataframe.columns):
        raise ValueError(f"Column number out of range: {column_index + 1}")

    search_rows: list[SearchRow] = []
    column_name = dataframe.columns[column_index]
    for dataframe_index, value in dataframe[column_name].items():
        keyword = "" if pd.isna(value) else str(value).strip()
        if keyword:
            search_rows.append(
                SearchRow(
                    row_number=int(dataframe_index) + resolved_header_row + 1,
                    keyword=keyword,
                )
            )

    return search_rows


def _read_csv(path: Path, **kwargs: Any) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return pd.read_csv(path, encoding=encoding, **kwargs)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, encoding="utf-8-sig", **kwargs)


def _format_header(value: Any, index: int) -> str:
    header = "" if value is None else str(value).strip()
    return header or f"Column {index + 1}"


def _score_header_rows(rows: Any) -> int:
    best_row = 1
    best_score = -1

    for zero_based_index, row in enumerate(rows):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if not values:
            continue

        joined = " ".join(values).lower()
        keyword_score = sum(1 for keyword in HEADER_KEYWORDS if keyword.lower() in joined)
        score = len(values) + keyword_score * 5
        if "bom of" in joined:
            score -= 10
        if score > best_score:
            best_row = zero_based_index + 1
            best_score = score

    return best_row

from __future__ import annotations

import csv
import json
import os
import tempfile
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def fill_xlsx_data(
    original_path: str | Path,
    output_path: str | Path,
    row_results: dict[int, dict[str, Any]],
    output_fields: list[str],
    sheet_name: str | None = None,
    header_row: int = 1,
    preserve_styles: bool = True,
) -> Path:
    original = Path(original_path)
    output = _prepare_output_path(original, output_path)
    workbook = load_workbook(original_path)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        start_col = worksheet.max_column + 1

        source_col = max(1, start_col - 1)
        source_width = worksheet.column_dimensions[get_column_letter(source_col)].width
        header_source = worksheet.cell(row=header_row, column=source_col)
        for offset, field_name in enumerate(output_fields):
            column_index = start_col + offset
            cell = worksheet.cell(row=header_row, column=start_col + offset, value=field_name)
            if preserve_styles:
                _copy_cell_style(header_source, cell)
            if source_width:
                worksheet.column_dimensions[get_column_letter(column_index)].width = source_width

        for row_index, values in row_results.items():
            source = worksheet.cell(row=row_index, column=source_col) if preserve_styles else None
            for offset, field_name in enumerate(output_fields):
                target = worksheet.cell(
                    row=row_index,
                    column=start_col + offset,
                    value=_format_cell_value(values.get(field_name)),
                )
                if source is not None:
                    _copy_cell_style(source, target)

        temp_path = _temporary_path(output)
        try:
            workbook.save(temp_path)
            os.replace(temp_path, output)
        finally:
            if temp_path.exists():
                temp_path.unlink()
    finally:
        workbook.close()
    return output


def write_csv_data(
    original_path: str | Path,
    output_path: str | Path,
    row_results: dict[int, dict[str, Any]],
    output_fields: list[str],
    header_row: int = 1,
) -> Path:
    original = Path(original_path)
    output = _prepare_output_path(original, output_path)

    reader = _read_csv_rows(original)

    if not reader:
        raise ValueError("CSV file is empty")
    if header_row < 1 or header_row > len(reader):
        raise ValueError(f"CSV header row out of range: {header_row}")

    header_index = header_row - 1
    reader[header_index].extend(output_fields)
    for row_number, row in enumerate(reader[header_index + 1 :], start=header_row + 1):
        values = row_results.get(row_number, {})
        row.extend(str(values.get(field_name, "")) for field_name in output_fields)

    temp_path = _temporary_path(output)
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as target_file:
            writer = csv.writer(target_file)
            writer.writerows(reader)
        os.replace(temp_path, output)
    finally:
        if temp_path.exists():
            temp_path.unlink()

    return output


def _prepare_output_path(original_path: Path, output_path: str | Path) -> Path:
    output = Path(output_path)
    if original_path.resolve() == output.resolve():
        raise ValueError("Output path must be different from the input file")
    output.parent.mkdir(parents=True, exist_ok=True)
    return output


def _temporary_path(output_path: Path) -> Path:
    handle = tempfile.NamedTemporaryFile(
        delete=False,
        dir=output_path.parent,
        prefix=f".{output_path.stem}.",
        suffix=output_path.suffix or ".tmp",
    )
    handle.close()
    return Path(handle.name)


def _read_csv_rows(path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as source_file:
                return list(csv.reader(source_file))
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError("CSV file could not be read")


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _format_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
SAMPLE_DIR = ROOT / "samples"


def main() -> int:
    SAMPLE_DIR.mkdir(exist_ok=True)
    output_path = SAMPLE_DIR / "sample_bom.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "BOM"
    worksheet.append(["序号", "器件型号", "数量", "备注"])
    worksheet.append([1, "STM32F103C8T6-TR", 10, "MCU"])
    worksheet.append([2, "0603-10K-1%-CT", 100, "电阻"])
    worksheet.append([3, "AMS1117-3.3-TR", 20, "LDO"])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    worksheet.column_dimensions["B"].width = 24
    workbook.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from bom_tool.core.file_writer import fill_xlsx_data


def test_fill_xlsx_data_copies_neighbor_styles_and_serializes_complex_values(
    workspace_tmp_path: Path,
) -> None:
    input_path = workspace_tmp_path / "input.xlsx"
    output_path = workspace_tmp_path / "output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["型号", "数量"])
    worksheet.append(["STM32F103C8T6", 10])
    worksheet.column_dimensions["B"].width = 18
    worksheet["B1"].font = Font(bold=True, color="FFFFFF")
    worksheet["B1"].fill = PatternFill("solid", fgColor="366092")
    worksheet["B2"].fill = PatternFill("solid", fgColor="D9EAF7")
    workbook.save(input_path)

    fill_xlsx_data(
        input_path,
        output_path,
        {2: {"lcsc_price_breaks": [{"quantity": 1, "unit_price": 8.5}]}},
        ["lcsc_price_breaks"],
    )

    exported = load_workbook(output_path)
    sheet = exported.active

    assert sheet["C1"].value == "lcsc_price_breaks"
    assert sheet["C1"].font.bold is True
    assert sheet["C1"].fill.fgColor.rgb == "00366092"
    assert sheet["C2"].fill.fgColor.rgb == "00D9EAF7"
    assert sheet.column_dimensions["C"].width == 18
    assert sheet["C2"].value == '[{"quantity": 1, "unit_price": 8.5}]'


def test_fill_xlsx_data_rejects_overwriting_input(workspace_tmp_path: Path) -> None:
    input_path = workspace_tmp_path / "input.xlsx"

    workbook = Workbook()
    workbook.active.append(["型号"])
    workbook.save(input_path)

    with pytest.raises(ValueError, match="Output path"):
        fill_xlsx_data(input_path, input_path, {}, ["状态"])


def test_fill_xlsx_data_fast_mode_skips_data_style_copy(workspace_tmp_path: Path) -> None:
    input_path = workspace_tmp_path / "input.xlsx"
    output_path = workspace_tmp_path / "output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["型号", "数量"])
    worksheet.append(["STM32F103C8T6", 10])
    worksheet["B2"].fill = PatternFill("solid", fgColor="D9EAF7")
    workbook.save(input_path)

    fill_xlsx_data(
        input_path,
        output_path,
        {2: {"stock": 100}},
        ["stock"],
        preserve_styles=False,
    )

    exported = load_workbook(output_path)
    sheet = exported.active

    assert sheet["C2"].value == 100
    assert sheet["C2"].fill.fill_type is None

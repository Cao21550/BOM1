from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from bom_tool.adapters.mock_adapter import MockSupplierAdapter
from bom_tool.core.bom_pipeline import BomPipeline, BomPipelineConfig, write_query_task_table
from bom_tool.core.file_parser import read_preview
from bom_tool.db.cache_db import CacheDB
from bom_tool.models import SearchType


@pytest.mark.asyncio
async def test_pipeline_reads_queries_and_writes_xlsx(workspace_tmp_path: Path) -> None:
    input_path = workspace_tmp_path / "input.xlsx"
    output_path = workspace_tmp_path / "output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["序号", "型号", "数量"])
    worksheet.append([1, "STM32F103C8T6-TR", 10])
    worksheet.append([2, "ABC-CT", 20])
    workbook.save(input_path)

    pipeline = BomPipeline([MockSupplierAdapter()])
    result = await pipeline.run(
        BomPipelineConfig(
            input_path=input_path,
            output_path=output_path,
            search_column="型号",
            output_fields=["status", "mpn", "stock", "price_unit"],
        )
    )

    assert result.total_rows == 2
    assert output_path.exists()

    exported = load_workbook(output_path)
    exported_sheet = exported.active
    headers = [cell.value for cell in exported_sheet[1]]
    assert headers[-4:] == ["模拟商城_状态", "模拟商城_型号", "模拟商城_库存", "模拟商城_单价"]
    assert exported_sheet.cell(row=2, column=5).value == "STM32F103C8T6"
    assert exported_sheet.cell(row=3, column=5).value == "ABC"


@pytest.mark.asyncio
async def test_pipeline_detects_header_row_and_writes_result_headers_there(
    workspace_tmp_path: Path,
) -> None:
    input_path = workspace_tmp_path / "input_with_title.xlsx"
    output_path = workspace_tmp_path / "output_with_title.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["BOM of demo board"])
    worksheet.append([])
    worksheet.append(["序号", "器件型号", "数量"])
    worksheet.append([1, "STM32F103C8T6-TR", 10])
    workbook.save(input_path)

    preview = read_preview(input_path)
    assert preview.header_row == 3
    assert preview.headers[:3] == ["序号", "器件型号", "数量"]

    pipeline = BomPipeline([MockSupplierAdapter()])
    await pipeline.run(
        BomPipelineConfig(
            input_path=input_path,
            output_path=output_path,
            search_column="器件型号",
            output_fields=["status", "mpn"],
        )
    )

    exported = load_workbook(output_path)
    exported_sheet = exported.active
    assert exported_sheet.cell(row=3, column=4).value == "模拟商城_状态"
    assert exported_sheet.cell(row=4, column=5).value == "STM32F103C8T6"


@pytest.mark.asyncio
async def test_pipeline_default_headers_are_chinese_without_price_breaks_or_lead_time(
    workspace_tmp_path: Path,
) -> None:
    input_path = workspace_tmp_path / "input_default_fields.xlsx"
    output_path = workspace_tmp_path / "output_default_fields.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["型号"])
    worksheet.append(["ABC"])
    workbook.save(input_path)

    pipeline = BomPipeline([MockSupplierAdapter()])
    await pipeline.run(
        BomPipelineConfig(
            input_path=input_path,
            output_path=output_path,
            search_column="型号",
            output_fields=None,
            enable_cache=False,
        )
    )

    exported = load_workbook(output_path)
    headers = [cell.value for cell in exported.active[1]]

    assert "模拟商城_查询关键字" in headers
    assert "模拟商城_状态" in headers
    assert "模拟商城_库存" in headers
    assert "模拟商城_阶梯价" not in headers
    assert "模拟商城_交期" not in headers
    assert all(not str(header).startswith("mock_") for header in headers if header)


def test_pipeline_precheck_reports_cache_hits_and_writes_task_table(
    workspace_tmp_path: Path,
) -> None:
    input_path = workspace_tmp_path / "input_precheck.xlsx"
    task_table_path = workspace_tmp_path / "query_tasks.csv"
    cache_path = workspace_tmp_path / "cache.sqlite3"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["型号"])
    worksheet.append(["ABC"])
    worksheet.append(["XYZ"])
    workbook.save(input_path)

    cache = CacheDB(cache_path)
    try:
        cache.set(
            "mock",
            SearchType.MPN,
            "ABC",
            {
                "supplier": "mock",
                "query": "ABC",
                "search_type": "mpn",
                "status": "success",
                "mpn": "ABC",
            },
        )
    finally:
        cache.close()

    pipeline = BomPipeline([MockSupplierAdapter()])
    result = pipeline.precheck(
        BomPipelineConfig(
            input_path=input_path,
            output_path=workspace_tmp_path / "output.xlsx",
            search_column="型号",
            cache_path=cache_path,
        )
    )
    write_query_task_table(result.task_records, task_table_path)

    assert result.total_rows == 2
    assert result.unique_queries == 2
    assert result.cached_adapter_results == 1
    assert result.network_adapter_results == 1
    assert task_table_path.exists()
    assert "needs_network" in task_table_path.read_text(encoding="utf-8-sig")
